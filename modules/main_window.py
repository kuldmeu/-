import os
from datetime import datetime
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QTableView, QPushButton, QLineEdit, QComboBox,
                             QLabel, QMenuBar, QMenu, QAction, QStatusBar,
                             QToolBar, QMessageBox, QHeaderView, QTabWidget,
                             QInputDialog, QFontDialog, QColorDialog, QFileDialog)
from PyQt5.QtCore import Qt, QSettings
from PyQt5.QtGui import QFont, QPalette, QColor
from .models import CounterpartyTableModel
from .dialogs import (AddCounterpartyDialog, ViewCounterpartyDialog,
                      CategoryManagerDialog, ServiceLocationManagerDialog,
                      SectionManagerDialog, StreetManagerDialog,
                      ReportDialog, MassOperationsDialog, SortDialog,
                      AboutDialog)
from .auth import UserManagementDialog, ChangePasswordDialog  # Импорт из auth.py
from .sort_dialog import SortDialog


class MainWindow(QMainWindow):
    def __init__(self, db_manager, user_data, settings):
        super().__init__()
        self.db_manager = db_manager
        self.user_data = user_data
        self.settings = settings
        self.table_model = None

        self.init_ui()
        self.load_settings()

    def init_ui(self):
        """Инициализация пользовательского интерфейса"""
        self.setWindowTitle("База данных контрагентов")
        self.setGeometry(100, 100, 1400, 800)

        # Создание центрального виджета
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Создание меню
        self.create_menu_bar()

        # Создание панели инструментов
        self.create_toolbar()

        # Создание панели поиска и фильтрации
        self.create_search_panel(layout)

        # Создание таблицы
        self.create_table(layout)

        # Создание информационной панели
        self.create_info_panel(layout)

        # Создание статусной строки
        self.create_status_bar()

        # Загрузка данных
        self.load_data()

    def create_menu_bar(self):
        """Создание меню"""
        menubar = self.menuBar()

        # Меню Файл
        file_menu = menubar.addMenu('Файл')

        backup_action = QAction('Резервное копирование', self)
        backup_action.triggered.connect(self.backup_database)
        file_menu.addAction(backup_action)

        restore_action = QAction('Восстановление из резервной копии', self)
        restore_action.triggered.connect(self.restore_database)
        file_menu.addAction(restore_action)

        export_action = QAction('Экспорт в Excel', self)
        export_action.triggered.connect(self.export_to_excel)
        file_menu.addAction(export_action)

        file_menu.addSeparator()

        exit_action = QAction('Выход', self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Меню Действия
        actions_menu = menubar.addMenu('Действия')

        add_action = QAction('Добавить контрагент', self)
        add_action.triggered.connect(self.add_counterparty)
        actions_menu.addAction(add_action)

        edit_action = QAction('Редактировать контрагент', self)
        edit_action.triggered.connect(self.edit_counterparty)
        actions_menu.addAction(edit_action)

        delete_action = QAction('Удалить контрагент', self)
        delete_action.triggered.connect(self.delete_counterparty)
        actions_menu.addAction(delete_action)

        refresh_action = QAction('Обновить данные', self)
        refresh_action.triggered.connect(self.load_data)
        actions_menu.addAction(refresh_action)

        # Меню Пользователь
        user_menu = menubar.addMenu('Пользователь')

        change_password_action = QAction('Сменить пароль', self)
        change_password_action.triggered.connect(self.change_password)
        user_menu.addAction(change_password_action)

        if self.user_data['role'] == 'admin':
            manage_users_action = QAction('Управление пользователями', self)
            manage_users_action.triggered.connect(self.manage_users)
            user_menu.addAction(manage_users_action)

        # Меню Отчет
        report_menu = menubar.addMenu('Отчет')

        template_report_action = QAction('Отчет по шаблону', self)
        template_report_action.triggered.connect(self.template_report)
        report_menu.addAction(template_report_action)

        summary_report_action = QAction('Сводный отчет', self)
        summary_report_action.triggered.connect(self.summary_report)
        report_menu.addAction(summary_report_action)

        # Меню База данных
        db_menu = menubar.addMenu('База данных')

        add_category_action = QAction('Добавить Категорию', self)
        add_category_action.triggered.connect(self.add_category)
        db_menu.addAction(add_category_action)

        add_section_action = QAction('Добавить Участок', self)
        add_section_action.triggered.connect(self.add_section)
        db_menu.addAction(add_section_action)

        add_service_location_action = QAction('Добавить Место оказание услуг', self)
        add_service_location_action.triggered.connect(self.add_service_location)
        db_menu.addAction(add_service_location_action)

        add_street_action = QAction('Добавить Улицу', self)
        add_street_action.triggered.connect(self.add_street)
        db_menu.addAction(add_street_action)

        # Меню Массовое изменение
        mass_menu = menubar.addMenu('Массовое изменение')

        mass_termination_action = QAction('Массовое расторжение договора', self)
        mass_termination_action.triggered.connect(self.mass_termination)
        mass_menu.addAction(mass_termination_action)

        mass_update_action = QAction('Массовое изменение показаний', self)
        mass_update_action.triggered.connect(self.mass_update_readings)
        mass_menu.addAction(mass_update_action)

        # Меню Вид
        view_menu = menubar.addMenu('Вид')

        change_theme_action = QAction('Изменить тему', self)
        change_theme_action.triggered.connect(self.change_theme)
        view_menu.addAction(change_theme_action)

        change_font_action = QAction('Изменить шрифт', self)
        change_font_action.triggered.connect(self.change_font)
        view_menu.addAction(change_font_action)

        change_layout_action = QAction('Настроить расположение', self)
        change_layout_action.triggered.connect(self.change_layout)
        view_menu.addAction(change_layout_action)

        # Меню Справка
        help_menu = menubar.addMenu('Справка')

        about_action = QAction('О программе', self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def create_toolbar(self):
        """Создание панели инструментов"""
        toolbar = QToolBar("Основные инструменты")
        self.addToolBar(toolbar)

        add_btn = QPushButton("Добавить контрагент")
        add_btn.clicked.connect(self.add_counterparty)
        toolbar.addWidget(add_btn)

        export_btn = QPushButton("Экспорт в Excel")
        export_btn.clicked.connect(self.export_to_excel)
        toolbar.addWidget(export_btn)

        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_data)
        toolbar.addWidget(refresh_btn)

    def create_search_panel(self, layout):
        """Создание панели поиска и фильтрации"""
        search_layout = QHBoxLayout()

        # Поле поиска по ID
        self.id_search = QLineEdit()
        self.id_search.setPlaceholderText("Поиск по ID")
        self.id_search.textChanged.connect(self.apply_filters)
        search_layout.addWidget(QLabel("ID:"))
        search_layout.addWidget(self.id_search)

        # Комбобокс категорий
        self.category_combo = QComboBox()
        self.category_combo.addItem("Все категории", None)
        self.load_categories()
        self.category_combo.currentIndexChanged.connect(self.apply_filters)
        search_layout.addWidget(QLabel("Категория:"))
        search_layout.addWidget(self.category_combo)

        # Поле поиска по ИИН/БИН
        self.iin_search = QLineEdit()
        self.iin_search.setPlaceholderText("Поиск по ИИН/БИН")
        self.iin_search.textChanged.connect(self.apply_filters)
        search_layout.addWidget(QLabel("ИИН/БИН:"))
        search_layout.addWidget(self.iin_search)

        # Кнопка сортировки
        sort_btn = QPushButton("Сортировка")
        sort_btn.clicked.connect(self.show_sort_dialog)
        search_layout.addWidget(sort_btn)

        layout.addLayout(search_layout)

    def create_table(self, layout):
        """Создание таблицы для отображения данных"""
        self.table_view = QTableView()
        self.table_model = CounterpartyTableModel([], self.db_manager)
        self.table_view.setModel(self.table_model)

        # Настройка внешнего вида таблицы
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.doubleClicked.connect(self.view_counterparty)

        # Включение контекстного меню для управления столбцами
        self.table_view.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.horizontalHeader().customContextMenuRequested.connect(self.show_column_menu)

        layout.addWidget(self.table_view)

    def show_column_menu(self, position):
        """Контекстное меню для управления столбцами"""
        menu = QMenu(self)

        for i in range(self.table_model.columnCount()):
            column_name = self.table_model.headerData(i, Qt.Horizontal)
            action = QAction(column_name, self)
            action.setCheckable(True)
            action.setChecked(not self.table_view.isColumnHidden(i))
            action.triggered.connect(lambda checked, col=i: self.toggle_column_visibility(col, checked))
            menu.addAction(action)

        menu.exec_(self.table_view.horizontalHeader().mapToGlobal(position))

    def toggle_column_visibility(self, column, visible):
        """Переключение видимости столбца"""
        self.table_view.setColumnHidden(column, not visible)

    def create_info_panel(self, layout):
        """Создание информационной панели"""
        info_layout = QHBoxLayout()

        self.total_contracts_label = QLabel("Всего договоров: 0")
        self.active_contracts_label = QLabel("Действующие: 0")
        self.terminated_contracts_label = QLabel("Расторгнутые: 0")
        self.available_contracts_label = QLabel("Есть в наличии: 0")
        self.not_available_contracts_label = QLabel("Нет в наличии: 0")

        info_layout.addWidget(self.total_contracts_label)
        info_layout.addWidget(self.active_contracts_label)
        info_layout.addWidget(self.terminated_contracts_label)
        info_layout.addWidget(self.available_contracts_label)
        info_layout.addWidget(self.not_available_contracts_label)
        info_layout.addStretch()

        layout.addLayout(info_layout)

    def create_status_bar(self):
        """Создание статусной строки"""
        status_bar = QStatusBar()
        self.setStatusBar(status_bar)
        status_bar.showMessage(f"Вход выполнен: {self.user_data['username']} ({self.user_data['role']})")

    def load_categories(self):
        """Загрузка категорий в комбобокс"""
        try:
            categories = self.db_manager.get_categories()
            for category in categories:
                self.category_combo.addItem(category['name'], category['id'])
        except Exception as e:
            print(f"Ошибка загрузки категорий: {e}")

    def load_data(self):
        """Загрузка данных в таблицу"""
        filters = {}

        # Применение фильтров
        if self.id_search.text():
            filters['id'] = self.id_search.text()

        if self.category_combo.currentData():
            filters['category_id'] = self.category_combo.currentData()

        if self.iin_search.text():
            filters['iin_bin'] = self.iin_search.text()

        data = self.db_manager.get_counterparties(filters)
        self.table_model.update_data(data)
        self.update_info_panel()

    def apply_filters(self):
        """Применение фильтров к данным"""
        self.load_data()

    def update_info_panel(self):
        """Обновление информационной панели"""
        total = len(self.table_model.counterparty_data)
        active = len([x for x in self.table_model.counterparty_data if x.get('contract_status') == 'active'])
        terminated = len([x for x in self.table_model.counterparty_data if x.get('contract_status') == 'terminated'])
        available = len(
            [x for x in self.table_model.counterparty_data if x.get('contract_availability') == 'Есть в наличии'])
        not_available = len(
            [x for x in self.table_model.counterparty_data if x.get('contract_availability') == 'Нет в наличии'])

        self.total_contracts_label.setText(f"Всего договоров: {total}")
        self.active_contracts_label.setText(f"Действующие: {active}")
        self.terminated_contracts_label.setText(f"Расторгнутые: {terminated}")
        self.available_contracts_label.setText(f"Есть в наличии: {available}")
        self.not_available_contracts_label.setText(f"Нет в наличии: {not_available}")

    def add_counterparty(self):
        """Добавление нового контрагента"""
        dialog = AddCounterpartyDialog(self.db_manager, self.user_data['id'], self)
        if dialog.exec_():
            self.load_data()

    def edit_counterparty(self):
        """Редактирование контрагента"""
        selection = self.table_view.selectionModel().selectedRows()
        if not selection:
            QMessageBox.warning(self, "Предупреждение", "Выберите контрагента для редактирования")
            return

        index = selection[0]
        # ИСПРАВЛЕНИЕ: используем свойство counterparty_data вместо data
        counterparty_data = self.table_model.counterparty_data[index.row()]
        dialog = AddCounterpartyDialog(self.db_manager, self.user_data['id'], self, counterparty_data)
        if dialog.exec_():
            self.load_data()

    def delete_counterparty(self):
        """Удаление контрагента"""
        selection = self.table_view.selectionModel().selectedRows()
        if not selection:
            QMessageBox.warning(self, "Предупреждение", "Выберите контрагента для удаления")
            return

        index = selection[0]
        counterparty_id = self.table_model.counterparty_data[index.row()]['id']  # ИСПРАВЛЕНО
        counterparty_name = self.table_model.counterparty_data[index.row()]['name']  # ИСПРАВЛЕНО

        reply = QMessageBox.question(self, "Подтверждение",
                                     f"Вы уверены, что хотите удалить контрагента '{counterparty_name}'?",
                                     QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            try:
                cursor = self.db_manager.conn.cursor()
                cursor.execute("DELETE FROM counterparties WHERE id = ?", (counterparty_id,))
                self.db_manager.conn.commit()
                self.load_data()
                QMessageBox.information(self, "Успех", "Контрагент удален")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка удаления: {e}")

    def view_counterparty(self, index):
        """Просмотр контрагента по двойному клику"""
        counterparty_data = self.table_model.counterparty_data[index.row()]  # ИСПРАВЛЕНО
        dialog = ViewCounterpartyDialog(counterparty_data['id'], self.db_manager, self.user_data, self)
        dialog.exec_()
        self.load_data()

    def export_to_excel(self):
        """Экспорт данных в Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            file_path, _ = QFileDialog.getSaveFileName(self, "Экспорт в Excel",
                                                       "контрагенты.xlsx", "Excel Files (*.xlsx)")
            if file_path:
                wb = Workbook()
                ws = wb.active
                ws.title = "Контрагенты"

                # Заголовки
                headers = []
                for col in range(self.table_model.columnCount()):
                    if not self.table_view.isColumnHidden(col):
                        headers.append(self.table_model.headerData(col, Qt.Horizontal))

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Данные - ИСПРАВЛЕНО
                for row, counterparty in enumerate(self.table_model.counterparty_data, 2):
                    col_index = 1
                    for col in range(self.table_model.columnCount()):
                        if not self.table_view.isColumnHidden(col):
                            index = self.table_model.index(row - 2, col)
                            value = self.table_model.data(index, Qt.DisplayRole)
                            ws.cell(row=row, column=col_index, value=value)
                            col_index += 1

                wb.save(file_path)
                QMessageBox.information(self, "Успех", "Данные экспортированы в Excel")

        except ImportError:
            QMessageBox.critical(self, "Ошибка", "Библиотека openpyxl не установлена")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {e}")

    def export_to_excel_with_volumes(self):
        """Экспорт в Excel с плановыми и фактическими объемами"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            import os

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Экспорт в Excel с объемами",
                f"контрагенты_с_объемами_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)")

            if not file_path:
                return

            wb = Workbook()

            # Лист с основными данными
            ws_main = wb.active
            ws_main.title = "Контрагенты"

            # Заголовки основных данных
            headers = []
            for col in range(self.table_model.columnCount()):
                if not self.table_view.isColumnHidden(col):
                    headers.append(self.table_model.headerData(col, Qt.Horizontal))

            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Данные
            for row, counterparty in enumerate(self.table_model.counterparty_data, 2):
                col_index = 1
                for col in range(self.table_model.columnCount()):
                    if not self.table_view.isColumnHidden(col):
                        index = self.table_model.index(row - 2, col)
                        value = self.table_model.data(index, Qt.DisplayRole)
                        ws_main.cell(row=row, column=col_index, value=value)
                        col_index += 1

            # Лист с объемами
            ws_volumes = wb.create_sheet("Объемы")

            # Заголовки для объемов
            volume_headers = ['Контрагент', 'Вид услуг', 'Тип услуги', 'Год', 'Месяц', 'Плановый объем',
                              'Фактический объем']
            for col, header in enumerate(volume_headers, 1):
                cell = ws_volumes.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Получение данных об объемах
            cursor = self.db_manager.conn.cursor()
            cursor.execute('''
                SELECT c.name as counterparty_name, pv.service_type, pv.subservice_type, pv.year,
                       pv.january, pv.february, pv.march, pv.april, pv.may, pv.june,
                       pv.july, pv.august, pv.september, pv.october, pv.november, pv.december
                FROM planned_volumes pv
                JOIN counterparties c ON pv.counterparty_id = c.id
            ''')

            volumes_data = cursor.fetchall()

            row_index = 2
            months = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                      'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']

            for volume in volumes_data:
                for month in months:
                    month_volume = volume.get(month, 0)
                    if month_volume > 0:
                        ws_volumes.cell(row=row_index, column=1, value=volume['counterparty_name'])
                        ws_volumes.cell(row=row_index, column=2, value=volume['service_type'])
                        ws_volumes.cell(row=row_index, column=3, value=volume['subservice_type'])
                        ws_volumes.cell(row=row_index, column=4, value=volume['year'])
                        ws_volumes.cell(row=row_index, column=5, value=month.capitalize())
                        ws_volumes.cell(row=row_index, column=6, value=month_volume)
                        # Фактические объемы можно добавить из другой таблицы
                        ws_volumes.cell(row=row_index, column=7, value=0)  # Заглушка
                        row_index += 1

            # Автоподбор ширины колонок
            for ws in [ws_main, ws_volumes]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Данные экспортированы в Excel:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта: {e}")

    def show_sort_dialog(self):
        """Показ диалога сортировки"""
        dialog = SortDialog(self.db_manager, self)
        result = dialog.exec_()

        if result == QDialog.Accepted:
            sort_params = dialog.get_sort_params()
            self.apply_sorting(sort_params)
        elif result == 2:  # Экспорт с объемами
            self.export_to_excel_with_volumes()

    def apply_sorting(self, sort_params):
        """Применение сортировки и фильтрации"""
        filters = {}

        # Применение фильтров
        if sort_params.get('category'):
            filters['category_id'] = sort_params['category']

        if sort_params.get('location'):
            filters['service_location_id'] = sort_params['location']

        if sort_params.get('status'):
            filters['contract_status'] = sort_params['status']

        if sort_params.get('availability'):
            filters['contract_availability'] = sort_params['availability']

        # Загрузка данных с фильтрами
        data = self.db_manager.get_counterparties(filters)
        self.table_model.update_data(data)

        # Применение сортировки
        if sort_params.get('sort_field'):
            # Реализация сортировки в модели
            pass

        self.update_info_panel()

    def load_settings(self):
        """Загрузка настроек приложения"""
        # Загрузка размеров и положения окна
        geometry = self.settings.value("geometry")
        if geometry:
            self.restoreGeometry(geometry)

        # Загрузка состояния таблицы
        header_state = self.settings.value("table_header_state")
        if header_state:
            self.table_view.horizontalHeader().restoreState(header_state)

        # Загрузка темы
        theme = self.settings.value("theme", "light")
        from .styles import AppStyles
        if theme == "dark":
            self.setPalette(AppStyles.get_dark_theme())
            self.setStyleSheet(AppStyles.get_stylesheet("dark"))
        elif theme == "blue":
            self.setPalette(AppStyles.get_blue_theme())
            self.setStyleSheet(AppStyles.get_stylesheet("blue"))
        else:
            self.setPalette(AppStyles.get_light_theme())
            self.setStyleSheet(AppStyles.get_stylesheet("light"))

        # Загрузка шрифта
        font_family = self.settings.value("font_family", "Arial")
        font_size = int(self.settings.value("font_size", 9))
        font = QFont(font_family, font_size)
        self.setFont(font)

    def closeEvent(self, event):
        """Событие закрытия приложения"""
        # Сохранение настроек
        self.settings.setValue("geometry", self.saveGeometry())
        self.settings.setValue("table_header_state", self.table_view.horizontalHeader().saveState())
        event.accept()

    def backup_database(self):
        """Резервное копирование базы данных"""
        backup_path, _ = QFileDialog.getSaveFileName(self, "Сохранить резервную копию",
                                                     "backup.db", "Database Files (*.db)")
        if backup_path:
            if self.db_manager.backup_database(backup_path):
                QMessageBox.information(self, "Успех", "Резервная копия создана")
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось создать резервную копию")

    def restore_database(self):
        """Восстановление базы данных"""
        backup_path, _ = QFileDialog.getOpenFileName(self, "Выберите резервную копию",
                                                     "", "Database Files (*.db)")
        if backup_path:
            reply = QMessageBox.question(self, "Подтверждение",
                                         "Восстановление базы данных перезапишет текущие данные. Продолжить?",
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                if self.db_manager.restore_database(backup_path):
                    self.load_data()
                    QMessageBox.information(self, "Успех", "База данных восстановлена")
                else:
                    QMessageBox.critical(self, "Ошибка", "Не удалось восстановить базу данных")

    def change_password(self):
        """Смена пароля"""
        dialog = ChangePasswordDialog(self.db_manager, self.user_data['id'], self)
        dialog.exec_()

    def manage_users(self):
        """Управление пользователями"""
        if self.user_data['role'] == 'admin':
            dialog = UserManagementDialog(self.db_manager, self)
            dialog.exec_()
        else:
            QMessageBox.warning(self, "Ошибка", "Недостаточно прав для управления пользователями")

    def template_report(self):
        """Отчет по шаблону"""
        dialog = ReportDialog(self.db_manager, self)
        dialog.exec_()

    def summary_report(self):
        """Сводный отчет"""
        dialog = ReportDialog(self.db_manager, self)
        dialog.tabs.setCurrentIndex(1)  # Переход на вкладку сводного отчета
        dialog.exec_()

    def add_category(self):
        """Добавление категории"""
        dialog = CategoryManagerDialog(self.db_manager, self)
        dialog.exec_()
        self.load_categories()

    def add_section(self):
        """Добавление участка"""
        dialog = SectionManagerDialog(self.db_manager, self)
        dialog.exec_()

    def add_service_location(self):
        """Добавление места оказания услуг"""
        dialog = ServiceLocationManagerDialog(self.db_manager, self)
        dialog.exec_()

    def add_street(self):
        """Добавление улицы"""
        dialog = StreetManagerDialog(self.db_manager, self)
        dialog.exec_()

    def mass_termination(self):
        """Массовое расторжение договоров"""
        dialog = MassOperationsDialog(self.db_manager, self.user_data['id'], self)
        dialog.exec_()
        self.load_data()

    def mass_update_readings(self):
        """Массовое изменение показаний"""
        dialog = MassOperationsDialog(self.db_manager, self.user_data['id'], self)
        dialog.mass_termination_radio.setChecked(False)
        dialog.mass_update_radio.setChecked(True)
        dialog.exec_()
        self.load_data()

    def change_theme(self):
        """Изменение темы оформления"""
        from .styles import AppStyles

        themes = ["Светлая", "Темная", "Синяя"]
        theme, ok = QInputDialog.getItem(self, "Выбор темы", "Выберите тему оформления:", themes, 0, False)

        if ok:
            if theme == "Темная":
                self.setPalette(AppStyles.get_dark_theme())
                self.setStyleSheet(AppStyles.get_stylesheet("dark"))
                self.settings.setValue("theme", "dark")
            elif theme == "Синяя":
                self.setPalette(AppStyles.get_blue_theme())
                self.setStyleSheet(AppStyles.get_stylesheet("blue"))
                self.settings.setValue("theme", "blue")
            else:
                self.setPalette(AppStyles.get_light_theme())
                self.setStyleSheet(AppStyles.get_stylesheet("light"))
                self.settings.setValue("theme", "light")

    def change_font(self):
        """Изменение шрифта"""
        font, ok = QFontDialog.getFont()
        if ok:
            self.setFont(font)
            self.settings.setValue("font_family", font.family())
            self.settings.setValue("font_size", font.pointSize())

    def change_layout(self):
        """Изменение расположения элементов"""
        QMessageBox.information(self, "Настройка расположения",
                                "Перетаскивайте элементы для изменения расположения. Изменения сохранятся автоматически.")

    def show_about(self):
        """Показ информации о программе"""
        dialog = AboutDialog(self)
        dialog.exec_()

    def load_settings(self):
        """Загрузка настроек приложения"""
        # Загрузка размеров и положения окна
        geometry = self.settings.value("geometry")
        if geometry:
            self.restoreGeometry(geometry)

        # Загрузка состояния таблицы
        header_state = self.settings.value("table_header_state")
        if header_state:
            self.table_view.horizontalHeader().restoreState(header_state)

        # Загрузка темы
        theme = self.settings.value("theme", "light")
        from .styles import AppStyles
        if theme == "dark":
            self.setPalette(AppStyles.get_dark_theme())
            self.setStyleSheet(AppStyles.get_stylesheet("dark"))
        elif theme == "blue":
            self.setPalette(AppStyles.get_blue_theme())
            self.setStyleSheet(AppStyles.get_stylesheet("blue"))
        else:
            self.setPalette(AppStyles.get_light_theme())
            self.setStyleSheet(AppStyles.get_stylesheet("light"))

        # Загрузка шрифта
        font_family = self.settings.value("font_family", "Arial")
        font_size = int(self.settings.value("font_size", 9))
        font = QFont(font_family, font_size)
        self.setFont(font)

    def closeEvent(self, event):
        """Событие закрытия приложения"""
        # Сохранение настроек
        self.settings.setValue("geometry", self.saveGeometry())
        self.settings.setValue("table_header_state", self.table_view.horizontalHeader().saveState())
        event.accept()