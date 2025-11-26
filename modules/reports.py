import os
from datetime import datetime
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QTabWidget,
                             QWidget, QLabel, QComboBox, QPushButton, 
                             QCheckBox, QGroupBox, QDateEdit, QTableWidget,
                             QTableWidgetItem, QFileDialog, QMessageBox,
                             QProgressDialog, QApplication)
from PyQt5.QtCore import QDate
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

class ReportDialog(QDialog):
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Генерация отчетов")
        self.setGeometry(100, 100, 800, 600)
        
        layout = QVBoxLayout()
        
        self.tabs = QTabWidget()
        
        # Вкладка "Отчет по шаблону"
        self.template_tab = QWidget()
        self.init_template_tab()
        self.tabs.addTab(self.template_tab, "Отчет по шаблону")
        
        # Вкладка "Сводный отчет"
        self.summary_tab = QWidget()
        self.init_summary_tab()
        self.tabs.addTab(self.summary_tab, "Сводный отчет")
        
        layout.addWidget(self.tabs)
        
        # Кнопки
        button_layout = QHBoxLayout()
        generate_btn = QPushButton("Сгенерировать отчет")
        generate_btn.clicked.connect(self.generate_report)
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(generate_btn)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def init_template_tab(self):
        layout = QVBoxLayout()
        
        # Выбор полей для отчета
        fields_group = QGroupBox("Выберите поля для отчета")
        fields_layout = QVBoxLayout()
        
        self.field_checkboxes = {}
        fields = [
            "ID", "ИИН/БИН", "Контрагент", "Место оказание услуг", "Участок",
            "Улица", "Дом", "Кв", "Объект", "Телефон", "Пред.показание",
            "Тек.показание", "Расход", "Номер автотранспорта", "Плановый объем",
            "Фактический объем", "Категория"
        ]
        
        # Размещение чекбоксов в 3 колонки
        columns_layout = QHBoxLayout()
        
        for i in range(3):
            column_layout = QVBoxLayout()
            start_idx = i * 6
            end_idx = min((i + 1) * 6, len(fields))
            
            for j in range(start_idx, end_idx):
                checkbox = QCheckBox(fields[j])
                self.field_checkboxes[fields[j]] = checkbox
                column_layout.addWidget(checkbox)
            
            column_layout.addStretch()
            columns_layout.addLayout(column_layout)
        
        fields_layout.addLayout(columns_layout)
        fields_group.setLayout(fields_layout)
        layout.addWidget(fields_group)
        
        self.template_tab.setLayout(layout)
    
    def init_summary_tab(self):
        layout = QVBoxLayout()
        
        # Параметры сводного отчета
        params_group = QGroupBox("Параметры отчета")
        params_layout = QGridLayout()
        
        params_layout.addWidget(QLabel("Категория:"), 0, 0)
        self.summary_category_combo = QComboBox()
        self.summary_category_combo.addItem("Все категории", None)
        params_layout.addWidget(self.summary_category_combo, 0, 1)
        
        params_layout.addWidget(QLabel("Место оказания услуг:"), 1, 0)
        self.summary_location_combo = QComboBox()
        self.summary_location_combo.addItem("Все места", None)
        params_layout.addWidget(self.summary_location_combo, 1, 1)
        
        params_layout.addWidget(QLabel("Участок:"), 2, 0)
        self.summary_section_combo = QComboBox()
        self.summary_section_combo.addItem("Все участки", None)
        params_layout.addWidget(self.summary_section_combo, 2, 1)
        
        params_layout.addWidget(QLabel("Период с:"), 3, 0)
        self.period_from = QDateEdit()
        self.period_from.setDate(QDate.currentDate().addMonths(-1))
        params_layout.addWidget(self.period_from, 3, 1)
        
        params_layout.addWidget(QLabel("по:"), 3, 2)
        self.period_to = QDateEdit()
        self.period_to.setDate(QDate.currentDate())
        params_layout.addWidget(self.period_to, 3, 3)
        
        params_layout.addWidget(QLabel("Год:"), 4, 0)
        self.year_combo = QComboBox()
        current_year = datetime.now().year
        for year in range(current_year - 5, current_year + 1):
            self.year_combo.addItem(str(year))
        self.year_combo.setCurrentText(str(current_year))
        params_layout.addWidget(self.year_combo, 4, 1)
        
        params_group.setLayout(params_layout)
        layout.addWidget(params_group)
        
        # Тип отчета
        report_type_group = QGroupBox("Тип отчета")
        report_type_layout = QVBoxLayout()
        
        self.summary_type_combo = QComboBox()
        self.summary_type_combo.addItems([
            "Отчет по договорам",
            "Отчет по плановым объемам", 
            "Отчет по фактическим объемам",
            "Сравнительный анализ"
        ])
        report_type_layout.addWidget(self.summary_type_combo)
        
        report_type_group.setLayout(report_type_layout)
        layout.addWidget(report_type_group)
        
        layout.addStretch()
        
        self.summary_tab.setLayout(layout)
        self.load_summary_filters()
    
    def load_summary_filters(self):
        """Загрузка фильтров для сводного отчета"""
        try:
            # Категории
            categories = self.db_manager.get_categories()
            for category in categories:
                self.summary_category_combo.addItem(category['name'], category['id'])
            
            # Места оказания услуг
            locations = self.db_manager.get_service_locations()
            for location in locations:
                self.summary_location_combo.addItem(location['name'], location['id'])
                
        except Exception as e:
            print(f"Ошибка загрузки фильтров: {e}")
    
    def generate_report(self):
        """Генерация отчета"""
        current_tab = self.tabs.currentIndex()
        
        if current_tab == 0:  # Отчет по шаблону
            self.generate_template_report()
        elif current_tab == 1:  # Сводный отчет
            self.generate_summary_report()
    
    def generate_template_report(self):
        """Генерация отчета по шаблону"""
        selected_fields = [field for field, checkbox in self.field_checkboxes.items() 
                          if checkbox.isChecked()]
        
        if not selected_fields:
            QMessageBox.warning(self, "Ошибка", "Выберите хотя бы одно поле для отчета")
            return
        
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Сохранить отчет", 
                f"отчет_по_шаблону_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            progress = QProgressDialog("Генерация отчета...", "Отмена", 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            # Получение данных
            progress.setValue(10)
            data = self.db_manager.get_counterparties()
            
            # Создание книги Excel
            progress.setValue(30)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет по шаблону"
            
            # Заголовки
            for col, field in enumerate(selected_fields, 1):
                cell = ws.cell(row=1, column=col, value=field)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Соответствие полей данным
            field_mapping = {
                "ID": "custom_id",
                "ИИН/БИН": "iin_bin", 
                "Контрагент": "name",
                "Место оказание услуг": "service_location_name",
                "Участок": "section_name",
                "Улица": "street_name",
                "Дом": "house",
                "Кв": "apartment",
                "Категория": "category_name",
                "Телефон": "phone"
            }
            
            # Данные
            progress.setValue(50)
            for row, counterparty in enumerate(data, 2):
                for col, field in enumerate(selected_fields, 1):
                    db_field = field_mapping.get(field, field.lower())
                    value = counterparty.get(db_field, '')
                    ws.cell(row=row, column=col, value=value)
                
                if row % 50 == 0:
                    progress.setValue(50 + (row * 40 // len(data)))
                    QApplication.processEvents()
            
            # Автоподбор ширины колонок
            progress.setValue(90)
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            progress.setValue(100)
            wb.save(file_path)
            progress.close()
            
            QMessageBox.information(self, "Успех", f"Отчет сохранен:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка генерации отчета: {e}")
    
    def generate_summary_report(self):
        """Генерация сводного отчета"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Сохранить сводный отчет",
                f"сводный_отчет_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Создание книги Excel
            wb = openpyxl.Workbook()
            
            # Статистика по договорам
            ws_stats = wb.active
            ws_stats.title = "Статистика договоров"
            
            # Получение данных
            filters = {}
            category_id = self.summary_category_combo.currentData()
            location_id = self.summary_location_combo.currentData()
            section_id = self.summary_section_combo.currentData()
            
            if category_id:
                filters['category_id'] = category_id
            if location_id:
                filters['service_location_id'] = location_id
            if section_id:
                filters['section_id'] = section_id
            
            data = self.db_manager.get_counterparties(filters)
            
            # Заголовки статистики
            headers = ['Показатель', 'Значение']
            for col, header in enumerate(headers, 1):
                cell = ws_stats.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Статистика
            stats = self.db_manager.get_statistics()
            statistics_data = [
                ['Всего договоров', stats.get('total', 0)],
                ['Действующие договоры', stats.get('active', 0)],
                ['Расторгнутые договоры', stats.get('terminated', 0)],
                ['Договоры в наличии', stats.get('available', 0)],
                ['Договоры не в наличии', stats.get('not_available', 0)]
            ]
            
            for row, (label, value) in enumerate(statistics_data, 2):
                ws_stats.cell(row=row, column=1, value=label)
                ws_stats.cell(row=row, column=2, value=value)
            
            # Создание диаграммы
            chart = PieChart()
            labels = Reference(ws_stats, min_col=1, min_row=2, max_row=6)
            data_ref = Reference(ws_stats, min_col=2, min_row=1, max_row=6)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels)
            chart.title = "Статистика договоров"
            ws_stats.add_chart(chart, "D2")
            
            # Лист с детализацией
            ws_detail = wb.create_sheet("Детализация")
            
            # Заголовки детализации
            detail_headers = ['ID', 'Контрагент', 'Категория', 'Место услуг', 'Статус', 'Наличие']
            for col, header in enumerate(detail_headers, 1):
                cell = ws_detail.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Данные детализации
            for row, counterparty in enumerate(data, 2):
                ws_detail.cell(row=row, column=1, value=counterparty.get('custom_id', ''))
                ws_detail.cell(row=row, column=2, value=counterparty.get('name', ''))
                ws_detail.cell(row=row, column=3, value=counterparty.get('category_name', ''))
                ws_detail.cell(row=row, column=4, value=counterparty.get('service_location_name', ''))
                ws_detail.cell(row=row, column=5, value=counterparty.get('contract_status', ''))
                ws_detail.cell(row=row, column=6, value=counterparty.get('contract_availability', ''))
            
            # Автоподбор ширины колонок
            for ws in [ws_stats, ws_detail]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Сводный отчет сохранен:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка генерации сводного отчета: {e}")

class MassOperationsDialog(QDialog):
    def __init__(self, db_manager, user_id, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.user_id = user_id
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Массовые операции")
        self.setGeometry(100, 100, 600, 500)
        
        layout = QVBoxLayout()
        
        # Выбор типа операции
        operation_group = QGroupBox("Тип операции")
        operation_layout = QVBoxLayout()
        
        self.mass_termination_radio = QRadioButton("Массовое расторжение договора")
        self.mass_update_radio = QRadioButton("Массовое изменение показаний")
        self.mass_termination_radio.setChecked(True)
        
        operation_layout.addWidget(self.mass_termination_radio)
        operation_layout.addWidget(self.mass_update_radio)
        
        operation_group.setLayout(operation_layout)
        layout.addWidget(operation_group)
        
        # Параметры фильтрации
        filter_group = QGroupBox("Параметры фильтрации")
        filter_layout = QGridLayout()
        
        filter_layout.addWidget(QLabel("Категория:"), 0, 0)
        self.mass_category_combo = QComboBox()
        self.mass_category_combo.addItem("Все категории", None)
        filter_layout.addWidget(self.mass_category_combo, 0, 1)
        
        filter_layout.addWidget(QLabel("Место оказания услуг:"), 1, 0)
        self.mass_location_combo = QComboBox()
        self.mass_location_combo.addItem("Все места", None)
        filter_layout.addWidget(self.mass_location_combo, 1, 1)
        
        filter_layout.addWidget(QLabel("Участок:"), 2, 0)
        self.mass_section_combo = QComboBox()
        self.mass_section_combo.addItem("Все участки", None)
        filter_layout.addWidget(self.mass_section_combo, 2, 1)
        
        filter_layout.addWidget(QLabel("Год:"), 3, 0)
        self.mass_year_combo = QComboBox()
        current_year = datetime.now().year
        for year in range(current_year - 5, current_year + 1):
            self.mass_year_combo.addItem(str(year))
        filter_layout.addWidget(self.mass_year_combo, 3, 1)
        
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        
        # Дополнительные параметры
        params_group = QGroupBox("Дополнительные параметры")
        params_layout = QVBoxLayout()
        
        self.reason_edit = QLineEdit()
        self.reason_edit.setPlaceholderText("Причина массовой операции...")
        params_layout.addWidget(QLabel("Причина:"))
        params_layout.addWidget(self.reason_edit)
        
        params_group.setLayout(params_layout)
        layout.addWidget(params_group)
        
        # Предварительный просмотр
        preview_group = QGroupBox("Предварительный просмотр")
        preview_layout = QVBoxLayout()
        
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(4)
        self.preview_table.setHorizontalHeaderLabels(['ID', 'Контрагент', 'Статус', 'Действие'])
        preview_layout.addWidget(self.preview_table)
        
        preview_btn = QPushButton("Обновить предпросмотр")
        preview_btn.clicked.connect(self.update_preview)
        preview_layout.addWidget(preview_btn)
        
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        # Кнопки
        button_layout = QHBoxLayout()
        apply_btn = QPushButton("Применить операцию")
        apply_btn.clicked.connect(self.apply_mass_operation)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(apply_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        self.load_filters()
        self.update_preview()
    
    def load_filters(self):
        """Загрузка фильтров"""
        try:
            categories = self.db_manager.get_categories()
            for category in categories:
                self.mass_category_combo.addItem(category['name'], category['id'])
            
            locations = self.db_manager.get_service_locations()
            for location in locations:
                self.mass_location_combo.addItem(location['name'], location['id'])
                
        except Exception as e:
            print(f"Ошибка загрузки фильтров: {e}")
    
    def update_preview(self):
        """Обновление предварительного просмотра"""
        try:
            filters = self.get_current_filters()
            data = self.db_manager.get_counterparties(filters)
            
            self.preview_table.setRowCount(len(data))
            for row, counterparty in enumerate(data):
                self.preview_table.setItem(row, 0, QTableWidgetItem(counterparty.get('custom_id', '')))
                self.preview_table.setItem(row, 1, QTableWidgetItem(counterparty.get('name', '')))
                self.preview_table.setItem(row, 2, QTableWidgetItem(counterparty.get('contract_status', '')))
                
                # Действие
                if self.mass_termination_radio.isChecked():
                    action = "Расторгнуть договор"
                else:
                    action = "Обновить показания"
                
                self.preview_table.setItem(row, 3, QTableWidgetItem(action))
                
        except Exception as e:
            print(f"Ошибка обновления предпросмотра: {e}")
    
    def get_current_filters(self):
        """Получение текущих фильтров"""
        filters = {}
        
        category_id = self.mass_category_combo.currentData()
        location_id = self.mass_location_combo.currentData()
        section_id = self.mass_section_combo.currentData()
        
        if category_id:
            filters['category_id'] = category_id
        if location_id:
            filters['service_location_id'] = location_id
        if section_id:
            filters['section_id'] = section_id
        
        return filters
    
    def apply_mass_operation(self):
        """Применение массовой операции"""
        if not self.reason_edit.text().strip():
            QMessageBox.warning(self, "Ошибка", "Укажите причину массовой операции")
            return
        
        reply = QMessageBox.question(
            self, "Подтверждение",
            "Вы уверены, что хотите выполнить массовую операцию?\n"
            "Это действие затронет все отфильтрованные записи.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                filters = self.get_current_filters()
                data = self.db_manager.get_counterparties(filters)
                
                if self.mass_termination_radio.isChecked():
                    self.mass_terminate_contracts(data)
                else:
                    self.mass_update_readings(data)
                
                QMessageBox.information(self, "Успех", "Массовая операция выполнена")
                self.accept()
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка выполнения операции: {e}")
    
    def mass_terminate_contracts(self, data):
        """Массовое расторжение договоров"""
        try:
            cursor = self.db_manager.conn.cursor()
            reason = self.reason_edit.text().strip()
            
            for counterparty in data:
                cursor.execute('''
                    UPDATE counterparties 
                    SET contract_status = 'terminated', updated_date = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (counterparty['id'],))
            
            self.db_manager.conn.commit()
            self.db_manager.data_updated.emit()
            
        except Exception as e:
            self.db_manager.conn.rollback()
            raise e
    
    def mass_update_readings(self, data):
        """Массовое обновление показаний"""
        # Реализация массового обновления показаний
        pass